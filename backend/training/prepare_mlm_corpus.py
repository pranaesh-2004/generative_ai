from openpyxl import load_workbook

FILES = [
    "D:\\student_feedback_genai\\backend\\data\\Feedback-Form-IV-Semester.xlsx",
    "D:\\student_feedback_genai\\backend\\data\\FEEDBACK-FORM-SEMESTER-VIII.xlsx"
]

OUTPUT_FILE = "feedback_corpus.txt"


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def is_valid_comment(text):
    if not text:
        return False
    if text.strip() == "-":
        return False
    return True


def extract_comments(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [clean_text(h) for h in rows[0]]
    comments = []

    for row in rows[1:]:
        for col_idx in range(1, len(headers)):  # skip Timestamp column
            text = clean_text(row[col_idx] if col_idx < len(row) else "")
            if is_valid_comment(text):
                comments.append(text)

    return comments


def main():
    all_comments = []

    for file_path in FILES:
        comments = extract_comments(file_path)
        all_comments.extend(comments)

    # remove exact duplicates, preserve order
    seen = set()
    unique_comments = []
    for comment in all_comments:
        if comment not in seen:
            seen.add(comment)
            unique_comments.append(comment)

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        for comment in unique_comments:
            f.write(comment.replace("\n", " ").strip() + "\n")

    print(f"Saved {len(unique_comments)} comments to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()