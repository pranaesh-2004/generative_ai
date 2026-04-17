import sys
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# Add backend folder to Python path
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from database.mongo import feedback_collection
from ai.sentiment import analyze_sentiment

DATA_DIR = BASE_DIR / "data"

FILES = [
    {
        "path": DATA_DIR / "Feedback-Form-IV-Semester.xlsx",
        "semester": 4
    },
    {
        "path": DATA_DIR / "FEEDBACK-FORM-SEMESTER-VIII.xlsx",
        "semester": 8
    }
]


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def is_valid_comment(text):
    if not text:
        return False
    if text.strip() == "-":
        return False
    return True


def parse_subject_and_faculty(header_text):
    """
    Example:
    22MDC41 Predictive Analytics(Dr.A.G.Aruna)
    ->
    subject_code = 22MDC41
    subject_name = Predictive Analytics
    faculty = Dr.A.G.Aruna
    """
    header_text = clean_text(header_text)

    match = re.match(r"^([A-Za-z0-9]+)\s+(.+?)(?:\((.*?)\))?$", header_text)
    if match:
        subject_code = clean_text(match.group(1))
        subject_name = clean_text(match.group(2))
        faculty = clean_text(match.group(3))
        return subject_code, subject_name, faculty

    return "", header_text, ""


def import_file(file_path, semester):
    if not file_path.exists():
        print(f"File not found: {file_path}")
        return 0

    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = [clean_text(h) for h in rows[0]]
    inserted_count = 0

    for row in rows[1:]:
        timestamp = row[0] if len(row) > 0 else None

        for col_index in range(1, len(headers)):
            header = headers[col_index]
            comment = clean_text(row[col_index] if col_index < len(row) else "")

            if not is_valid_comment(comment):
                continue

            subject_code, subject_name, faculty = parse_subject_and_faculty(header)
            sentiment_result = analyze_sentiment(comment)

            feedback_doc = {
                "semester": semester,
                "subject_code": subject_code,
                "subject": subject_name,
                "faculty": faculty,
                "comment": comment,
                "sentiment": sentiment_result["sentiment"],
                "confidence": sentiment_result["confidence"],
                "source_file": file_path.name,
                "created_at": timestamp if isinstance(timestamp, datetime) else datetime.utcnow()
            }

            feedback_collection.insert_one(feedback_doc)
            inserted_count += 1

    return inserted_count


def main():
    total_inserted = 0

    print("BASE_DIR:", BASE_DIR)
    print("DATA_DIR:", DATA_DIR)

    for file_info in FILES:
        count = import_file(file_info["path"], file_info["semester"])
        print(f"Inserted {count} records from {file_info['path'].name}")
        total_inserted += count

    print(f"Total inserted: {total_inserted}")


if __name__ == "__main__":
    main()